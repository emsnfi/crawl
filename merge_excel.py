
import pandas as pd
import xlrd as xlrd
import openpyxl as xl
import re
import time
from datetime import datetime
from xlrd import xldate_as_tuple
#import xlwt as xlwt
#data = pd.read_excel('/Users/emily/Desktop/巨量資料與金融科技實務期末報告檔案/1Q Beta_Monthly (1).xlsx',sheet_name='工作表5')
file_name="/Users/emily/Downloads/巨量資料與金融科技實務期末報告檔案-copy/1Q_Chg._Alpha_Monthly (1).xlsx"
data = xlrd.open_workbook(file_name)
sheet_merge = data.sheet_by_name("合併月報酬資料")

wb  = xl.load_workbook(file_name)
ws = wb.worksheets[3] #合併月報酬資料

s = '0123456789'
company = sheet_merge.col_values(1) 
company_s =[]
company_word=[]#只有文字的
company_id=[]

#取公司id
for i in range(len(company)):
    str = company[i]
    tmp = ''
    for j in range(len(str)):
        if str[j].isdigit():
            tmp += str[j]
    company_id.append(tmp)

#print(company_id)

#取公司 名字
for i in range(len(company)):
    company_s.append(company[i])

    company_word.append(re.sub('[0123456789]', '',company_s[i]))
    
    
#insert column in excel, and put data in it 要轉成def
#wsnew = ws.insert_cols(3)
#for index,row in enumerate(ws.rows):
 #   row[2].value = company_word[index]
#for index,row in enumerate(ws.rows):
 #   row[2].value = company_id[index]
#wb.save('/Users/emily/Desktop/巨量資料與金融科技實務期末報告檔案-2/1Q Beta_Monthly (1) .xlsx')
#https://blog.csdn.net/gcgaochuang/article/details/101222414?utm_medium=distribute.pc_relevant_t0.none-task-blog-BlogCommendFromMachineLearnPai2-1.nonecase&depth_1-utm_source=distribute.pc_relevant_t0.none-task-blog-BlogCommendFromMachineLearnPai2-1.nonecase
#合併檔案
ws_re = wb.worksheets[2] #return的工作表
sheet_return = data.sheet_by_name("return")
company_re = sheet_return.col_values(1) #return sheet的公司行
time_return = sheet_return.row_values(1)
time_merge = sheet_merge.row_values(0)

#時間處理 待處理
#for a in range(4,len(time_merge)):
    #date = datetime(*xldate_as_tuple(sCell, 0))
    #print(date.strftime(time_merge[a]))

start =-1
return_ro=[]
re=[]
result=[]

newfile="/Users/emily/Desktop/merge_al.xlsx"
wother = xl.load_workbook(newfile)
wother.create_sheet('merge')
we = wother.worksheets[1]
start=-1
for t in range(2,len(time_merge)):
    for tr in range(2,len(time_return)):
        if(time_return[tr] == time_merge[t]):
            
            if(start==-1):
                    start=tr;

#company_s
for i in range(len(company_s)):
    
    for j in range(len((company_re))):
        if company_s[i] == company_re[j]  : #公司名有對到
            #print(company_re[j])
            
            return_ro=sheet_return.row_values(j)
            re.extend(return_ro[start:])#把從2000的return資料放進re裡
            break;          
    print(return_ro[start:],end='\nchange')
    we.append(return_ro[start:])
    
#return_ro.clear()
    #wsnew = we.insert_rows(2)
    #print(re,end="end")
    #we.append(re)
    #print(re[1],end="以")
                        #print(start)
            #print(sheet_return.row_values(j),end="/n")
        #for c in range(4,len(time_merge)):
            #sheet_merge.cell_value(i,4) = re
            #row[i](c).value = re
            #ws.cell(row=i+1, column=c+1).value = 'tree'
         #   ws.write(i, c, re)
        
        #wb.save('/Users/emily/Desktop/巨量資料與金融科技實務期末報告檔案-2/1Q Beta_Monthly (1) .xlsx')
wother.save('/Users/emily/Desktop/merge_al.xlsx')
    
    
        
            
                        #print(time_merge[t])
                        #print(time_return[tr])
    #print(re)
    
            
        #ws.cell(row=i, column=c).value = re[c-4]
        #ws.write(i,c,re[c-4])
#把 return內容放入merge內
#print(result[3])


print("start")
print(start)                  
    






